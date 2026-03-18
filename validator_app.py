import streamlit as st
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import zipfile
import shutil
import tempfile
import os
import io
import xml.etree.ElementTree as ET
import pandas as pd

# ---------------------------------------------------------------------------
# Validation rule definitions (keyed by 0-based column index)
# ---------------------------------------------------------------------------
COLUMN_RULES = {
    0:  {"header": "Name",                  "type": "text",     "required": True},
    1:  {"header": "Fathers Name",          "type": "text",     "required": True},
    2:  {"header": "Grand fathers Name",    "type": "text",     "required": True},
    3:  {"header": "Gender",                "type": "dropdown", "required": True,
         "allowed": ["F", "M"]},
    4:  {"header": "Age",                   "type": "integer",  "required": True,
         "min": 18, "max": 80},
    5:  {"header": "Education Level",       "type": "dropdown", "required": True,
         "allowed": ["Diploma", "Bachelor", "Master", "Director", "Doctorate", "Professional"]},
    6:  {"header": "Field Of Study",        "type": "text",     "required": True},
    7:  {"header": "University/College",    "type": "text",     "required": True},
    8:  {"header": "Subject of Teaching",   "type": "dropdown", "required": True,
         "allowed": [
             "Management", "Math", "Chemistry", "Physics", "English",
             "Afan Oromo", "Amharic", "HPE", "Geography",
             "Civics / Citizenship", "History", "Biology", "ICT",
             "General Science", "SNE", "Social Studies",
             "Environmental Science", "Moral Education", "PVA",
             "HPE / In Amharic", "HPE / In Afan Oromo",
             "Math / In Amharic", "Math / In Afan Oromo",
             "Environmental / In Amharic", "Environmental / In Afan Oromo",
             "Moral / In Amharic", "Moral / In Afan Oromo",
             "PVA / In Amharic", "PVA / In Afan Oromo",
             "Principal / In Amharic", "Principal / In Afan Oromo",
             "Supervisor / In Amharic", "Supervisor / In Afan Oromo",
         ]},
    9:  {"header": "Job Title",             "type": "text",     "required": True},
    10: {"header": "Has Taken Educational Leadership Related Course PGDSL/PGDSLM",
         "type": "dropdown", "required": True, "allowed": ["Yes", "No"]},
    11: {"header": "Date of Employment",    "type": "date",     "required": True},
    12: {"header": "Career Ladder",         "type": "dropdown", "required": True,
         "allowed": ["Beginner", "Junior", "Higher", "Associate",
                      "Associate Lead", "Lead", "Senior Lead",
                      "Senior Lead Two", "Senior Lead Three"]},
    13: {"header": "School Ownership",      "type": "dropdown", "required": True,
         "allowed": ["Private", "Public", "Government", "Unknown"]},
    14: {"header": "School Name",           "type": "text",     "required": True},
    15: {"header": "Level Of The School",   "type": "dropdown", "required": True,
         "allowed": ["KG", "Primary School", "Primary and Middle School",
                      "Secondary School"]},
    16: {"header": "SUB-CITY",              "type": "dropdown", "required": True,
         "allowed": ["Bole", "Arada", "Gulale", "Lami Kura",
                      "Kolfe Karanio", "Nifas Silk Lafto", "Akaki Qality",
                      "Kirkos", "Adis Ketama", "Lideta", "Yeka"]},
    17: {"header": "Type of Licence Owned", "type": "dropdown", "required": False,
         "allowed": ["Temporary", "Entry level", "Full", "Permanent", "SAT"]},
    18: {"header": "Date of Licence Owned", "type": "date",     "required": False},
    19: {"header": "Mobile Number",         "type": "phone",    "required": True},
    20: {"header": "Disability",            "type": "dropdown", "required": True,
         "allowed": ["None", "HearingImpairment", "VisualImpairment",
                      "MobilityImpairment", "Autism", "Other"]},
    21: {"header": "Activity Type",         "type": "dropdown", "required": True,
         "allowed": ["Teacher", "Director", "ViceDirector", "Supervisor",
                      "Unknown"]},
    22: {"header": "Total Experience (Years)", "type": "number", "required": False},
    23: {"header": "Carrier number",        "type": "text",     "required": False},
}

TOTAL_COLUMNS = 24  # A–X
COL_LETTERS = [chr(ord("A") + i) for i in range(TOTAL_COLUMNS)]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def excel_serial_to_date(serial):
    """Convert an Excel date serial number to a Python datetime."""
    if isinstance(serial, (int, float)) and serial > 0:
        return datetime(1899, 12, 30) + timedelta(days=int(serial))
    return None


def normalize_value(value):
    """Return a stripped string or the value itself (for dates/numbers)."""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def is_empty(value):
    """Check if a value is empty/missing."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def is_row_empty(row):
    """Check if an entire row is empty (all cells None or blank)."""
    return all(is_empty(cell) for cell in row)


def validate_cell(value, rule):
    """Validate a single cell value against its rule. Returns error message or None."""
    raw = value
    value = normalize_value(value)
    rtype = rule["type"]
    required = rule["required"]
    header = rule["header"]

    # --- Handle empty / missing ---
    if is_empty(value):
        if required:
            return f"{header}: value is empty or missing (required)"
        else:
            return None  # optional and empty → OK

    # --- Type-specific validation ---
    if rtype == "text":
        if not isinstance(value, str):
            return f"{header}: expected text, got {type(value).__name__}"
        return None

    if rtype == "dropdown":
        str_val = str(value).strip()
        allowed = rule["allowed"]
        if str_val not in allowed:
            allowed_preview = ", ".join(allowed[:10])
            if len(allowed) > 10:
                allowed_preview += f" ... ({len(allowed)} total)"
            return f"{header}: value '{str_val}' is not allowed. Allowed: {allowed_preview}"
        return None

    if rtype == "integer":
        try:
            num = int(float(value))
        except (ValueError, TypeError):
            return f"{header}: expected an integer, got '{value}'"
        lo, hi = rule.get("min", 0), rule.get("max", 999)
        if num < lo or num > hi:
            return f"{header}: value {num} is out of range ({lo}–{hi})"
        return None

    if rtype == "date":
        if isinstance(raw, datetime):
            return None
        if isinstance(value, (int, float)):
            dt = excel_serial_to_date(value)
            if dt is None:
                return f"{header}: invalid date serial number '{value}'"
            return None
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    datetime.strptime(value, fmt)
                    return None
                except ValueError:
                    continue
            return f"{header}: '{value}' is not a valid date"
        return f"{header}: expected a date, got '{value}'"

    if rtype == "phone":
        digits = str(value).strip().replace(" ", "").replace("-", "")
        # If it looks like a float from Excel (e.g. 912345678.0), clean it
        if "." in digits:
            try:
                digits = str(int(float(digits)))
            except (ValueError, TypeError):
                pass
        if not digits.isdigit() or len(digits) != 9:
            return f"{header}: expected a 9-digit number, got '{value}'"
        return None

    if rtype == "number":
        try:
            num = float(value)
        except (ValueError, TypeError):
            return f"{header}: expected a number, got '{value}'"
        if num < 0:
            return f"{header}: value {num} must be non-negative"
        return None

    return None


def display_value(value, col_idx):
    """Return a display-friendly string for a cell value."""
    if value is None:
        return ""
    rule = COLUMN_RULES.get(col_idx)
    if rule and rule["type"] == "date":
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, (int, float)):
            dt = excel_serial_to_date(value)
            if dt:
                return dt.strftime("%Y-%m-%d")
    # Avoid displaying whole numbers as floats (e.g. 912345678.0 → 912345678)
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------

def _strip_data_validations(file_bytes):
    """Remove dataValidation elements from sheet XML to work around
    malformed data-validation entries that crash openpyxl."""
    tmp_dir = tempfile.mkdtemp()
    cleaned_path = os.path.join(tmp_dir, "cleaned.xlsx")
    try:
        with zipfile.ZipFile(file_bytes, "r") as zin:
            with zipfile.ZipFile(cleaned_path, "w") as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                        tree = ET.ElementTree(ET.fromstring(data))
                        root = tree.getroot()
                        for dv in root.findall(f"{{{ns}}}dataValidations"):
                            root.remove(dv)
                        data = ET.tostring(root, xml_declaration=True, encoding="UTF-8")
                    zout.writestr(item, data)
        return cleaned_path, tmp_dir
    except Exception:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise


def load_workbook_safe(file):
    """Load workbook, stripping malformed data validations if needed.
    Returns (workbook, tmp_dir_to_cleanup_or_None)."""
    tmp_dir = None
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception:
        if hasattr(file, "seek"):
            file.seek(0)
        cleaned_path, tmp_dir = _strip_data_validations(file)
        wb = openpyxl.load_workbook(cleaned_path, data_only=True)
    return wb, tmp_dir


def get_sheet_names(file):
    """Return list of sheet names from an Excel file."""
    wb, tmp_dir = load_workbook_safe(file)
    names = wb.sheetnames
    if tmp_dir:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    return names


def read_excel(file, sheet_name=None):
    """Read an uploaded Excel file and return (headers, rows).
    Each row is a list of raw cell values (length = TOTAL_COLUMNS).
    Empty rows are skipped.
    """
    wb, tmp_dir = load_workbook_safe(file)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        all_rows = list(ws.iter_rows(min_row=1, max_col=TOTAL_COLUMNS, values_only=True))
        if not all_rows:
            return [], []
        headers = [normalize_value(c) or "" for c in all_rows[0]]
        data_rows = []
        for row in all_rows[1:]:
            padded = list(row) + [None] * (TOTAL_COLUMNS - len(row))
            padded = padded[:TOTAL_COLUMNS]
            if is_row_empty(padded):
                continue  # skip completely empty rows
            data_rows.append(padded)
        return headers, data_rows
    finally:
        if tmp_dir:
            shutil.rmtree(tmp_dir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Validation engine
# ---------------------------------------------------------------------------

def validate_rows(rows, selected_columns):
    """Validate data rows. Returns dict: row_index -> list of (col_idx, error_msg)."""
    errors = {}
    for row_idx, row in enumerate(rows):
        row_errors = []
        for col_idx in selected_columns:
            rule = COLUMN_RULES.get(col_idx)
            if rule is None:
                continue
            err = validate_cell(row[col_idx], rule)
            if err:
                row_errors.append((col_idx, err))
        if row_errors:
            errors[row_idx] = row_errors
    return errors


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

def build_html_table(headers, rows, errors, selected_columns):
    """Build an HTML table with color-coded cells."""
    error_cells = set()
    for row_idx, errs in errors.items():
        for col_idx, _ in errs:
            error_cells.add((row_idx, col_idx))

    html = """<style>
    .vtable { border-collapse: collapse; width: 100%; font-size: 13px; }
    .vtable th { background: #262730; color: #fafafa; padding: 6px 8px;
                 border: 1px solid #444; text-align: left; white-space: nowrap; }
    .vtable td { padding: 5px 8px; border: 1px solid #444; color: #fafafa; }
    .cell-ok  { background-color: #1e3a2a; }
    .cell-err { background-color: #3a1e1e; }
    .cell-skip { background-color: #1e1e2e; }
    </style>"""
    html += '<div style="overflow-x:auto;"><table class="vtable"><thead><tr>'
    html += "<th>Row</th>"
    for ci in range(TOTAL_COLUMNS):
        html += f"<th>{headers[ci] if ci < len(headers) else ''}</th>"
    html += "</tr></thead><tbody>"

    for ri, row in enumerate(rows):
        html += "<tr>"
        html += f"<td><b>{ri + 2}</b></td>"
        for ci in range(TOTAL_COLUMNS):
            val = display_value(row[ci], ci)
            if ci not in selected_columns:
                cls = "cell-skip"
            elif (ri, ci) in error_cells:
                cls = "cell-err"
            else:
                cls = "cell-ok"
            safe_val = (val.replace("&", "&amp;").replace("<", "&lt;")
                           .replace(">", "&gt;").replace('"', "&quot;"))
            html += f'<td class="{cls}">{safe_val}</td>'
        html += "</tr>"

    html += "</tbody></table></div>"
    return html


def main():
    st.set_page_config(page_title="AP Excel Validator", layout="wide")

    # ---- Sidebar ----
    with st.sidebar:
        st.title("AP Excel Validator")
        st.markdown(
            "Upload a filled **AcademicPersonel** Excel file (`.xlsx`) "
            "to validate it against the template rules."
        )
        st.markdown("---")
        uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

    if uploaded_file is None:
        st.info("Upload an Excel file using the sidebar to get started.")
        return

    # ---- Detect sheets ----
    file_bytes = uploaded_file.read()
    sheet_names = get_sheet_names(io.BytesIO(file_bytes))

    selected_sheet = sheet_names[0]
    if len(sheet_names) > 1:
        st.subheader("Select Sheet")
        selected_sheet = st.selectbox(
            "This file has multiple sheets. Choose one to validate:",
            sheet_names,
        )

    # ---- Per-sheet persistence setup ----
    if "sheet_edits" not in st.session_state:
        st.session_state.sheet_edits = {}   # {sheet_name: rows}
    if "sheet_headers" not in st.session_state:
        st.session_state.sheet_headers = {} # {sheet_name: headers}
    if "sheet_errors" not in st.session_state:
        st.session_state.sheet_errors = {}  # {sheet_name: errors}
    if "sheet_validated_cols" not in st.session_state:
        st.session_state.sheet_validated_cols = {}  # {sheet_name: selected_columns}

    # ---- Read the selected sheet ----
    try:
        headers, rows = read_excel(io.BytesIO(file_bytes), sheet_name=selected_sheet)
    except Exception as e:
        st.error(f"Failed to read Excel file: {e}")
        return

    # Always store original headers per sheet
    st.session_state.sheet_headers[selected_sheet] = headers

    if not rows:
        st.warning("The selected sheet has no data rows (only a header or empty).")
        return

    # Show which sheets have been edited
    if len(sheet_names) > 1:
        edited_sheets = [s for s in sheet_names if s in st.session_state.sheet_edits]
        if edited_sheets:
            st.info(f"Sheets with saved edits: **{', '.join(edited_sheets)}**")

    sheet_label = f" (sheet: **{selected_sheet}**)" if len(sheet_names) > 1 else ""
    st.success(
        f"File loaded: **{uploaded_file.name}**{sheet_label} — "
        f"{len(rows)} data row(s) found (empty rows excluded)."
    )

    # ---- Column Selection ----
    st.subheader("Select Columns to Validate")

    # Use individual session-state keys for each checkbox so the toggle works
    CHECKBOX_KEYS = [f"chk_{i}" for i in range(TOTAL_COLUMNS)]

    # Initialize checkbox state on first run
    if "col_init_done" not in st.session_state:
        for key in CHECKBOX_KEYS:
            st.session_state[key] = True
        st.session_state.col_init_done = True

    # Select All / Deselect All — directly mutate the checkbox keys
    if st.button("Select All / Deselect All"):
        all_selected = all(st.session_state[k] for k in CHECKBOX_KEYS)
        new_val = not all_selected
        for key in CHECKBOX_KEYS:
            st.session_state[key] = new_val
        st.rerun()

    # Render checkboxes in a 4-column grid
    grid_cols = st.columns(4)
    for i in range(TOTAL_COLUMNS):
        rule = COLUMN_RULES.get(i, {})
        label = f"{COL_LETTERS[i]} - {rule.get('header', '?')}"
        with grid_cols[i % 4]:
            st.checkbox(label, key=CHECKBOX_KEYS[i])

    selected_columns = sorted(
        i for i in range(TOTAL_COLUMNS) if st.session_state[CHECKBOX_KEYS[i]]
    )

    # Run Validation button
    run_clicked = st.button("Run Validation", type="primary")

    # Check if this sheet already has saved validation results
    has_prior_results = selected_sheet in st.session_state.sheet_errors

    if not run_clicked and not has_prior_results:
        st.info("Select the columns you want to validate, then click **Run Validation**.")
        return

    # ---- Run validation ----
    if run_clicked:
        errors = validate_rows(rows, selected_columns)
        st.session_state.sheet_errors[selected_sheet] = errors
        st.session_state.sheet_validated_cols[selected_sheet] = selected_columns
        # Clear per-sheet edits on fresh validation
        st.session_state.sheet_edits.pop(selected_sheet, None)
    else:
        errors = st.session_state.sheet_errors[selected_sheet]
        selected_columns = st.session_state.sheet_validated_cols[selected_sheet]

    # Use edited rows for this sheet if available
    active_rows = st.session_state.sheet_edits.get(selected_sheet, rows)

    # ---- Summary Dashboard ----
    st.markdown("---")
    st.subheader("Validation Summary")

    total_rows = len(active_rows)
    total_error_count = sum(len(e) for e in errors.values())
    invalid_rows = len(errors)
    valid_rows = total_rows - invalid_rows

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Rows", total_rows)
    m2.metric("Total Errors", total_error_count)
    m3.metric("Valid Rows", valid_rows)
    m4.metric("Invalid Rows", invalid_rows)

    # Errors per column bar chart
    if total_error_count > 0:
        col_error_counts = defaultdict(int)
        for errs in errors.values():
            for col_idx, _ in errs:
                rule = COLUMN_RULES.get(col_idx, {})
                col_error_counts[f"{COL_LETTERS[col_idx]} - {rule.get('header', '?')}"] += 1
        chart_df = pd.DataFrame(
            list(col_error_counts.items()), columns=["Column", "Errors"]
        ).sort_values("Errors", ascending=False)
        st.bar_chart(chart_df.set_index("Column"))

    # ---- Tabs: Validation View / Edit & Fix ----
    st.markdown("---")
    tab_view, tab_edit = st.tabs(["Validation View", "Edit & Fix"])

    # -- Tab 1: Color-coded read-only table + error details --
    with tab_view:
        st.subheader("Data Table")
        html_table = build_html_table(headers, active_rows, errors, set(selected_columns))
        st.markdown(html_table, unsafe_allow_html=True)

        st.markdown("---")
        st.subheader("Error Details")

        show_errors_only = st.checkbox("Show only rows with errors", value=False)

        if total_error_count == 0:
            st.success("No validation errors found!")
        else:
            for row_idx in range(total_rows):
                if row_idx not in errors:
                    if show_errors_only:
                        continue
                    st.write(f"**Row {row_idx + 2}** — Valid")
                else:
                    row_errs = errors[row_idx]
                    with st.expander(f"Row {row_idx + 2} — {len(row_errs)} error(s)", expanded=False):
                        for col_idx, msg in row_errs:
                            st.error(msg)

    # -- Tab 2: Editable table + re-validate + download --
    with tab_edit:
        st.subheader("Edit Data")
        st.caption("Fix errors directly in the table below, then re-validate or download.")

        # Show compact error guide if there are errors
        if total_error_count > 0:
            with st.expander(f"**{total_error_count} error(s) to fix** — click to see where", expanded=True):
                for row_idx in sorted(errors.keys()):
                    row_errs = errors[row_idx]
                    cols_with_errors = [
                        f"**{COL_LETTERS[ci]}** ({COLUMN_RULES[ci]['header']}): {msg.split(': ', 1)[1]}"
                        for ci, msg in row_errs
                    ]
                    st.markdown(
                        f"Row **{row_idx + 2}** — " + " | ".join(cols_with_errors)
                    )

        # Build a DataFrame for editing with an Errors column
        col_names = [f"{COL_LETTERS[i]} - {headers[i]}" if i < len(headers) else COL_LETTERS[i]
                     for i in range(TOTAL_COLUMNS)]
        display_rows = []
        error_markers = []
        for ri, row in enumerate(active_rows):
            display_rows.append([display_value(row[ci], ci) for ci in range(TOTAL_COLUMNS)])
            if ri in errors:
                err_cols = [COL_LETTERS[ci] for ci, _ in errors[ri]]
                error_markers.append(f"{len(errors[ri])}  [{', '.join(err_cols)}]")
            else:
                error_markers.append("")

        edit_df = pd.DataFrame(display_rows, columns=col_names)
        edit_df.insert(0, "Errors", error_markers)
        edit_df.index = [i + 2 for i in range(len(edit_df))]
        edit_df.index.name = "Row"

        # Make the Errors column non-editable
        column_config = {"Errors": st.column_config.TextColumn("Errors", disabled=True)}
        edited_df = st.data_editor(
            edit_df, use_container_width=True, num_rows="dynamic",
            column_config=column_config,
        )

        col_a, col_b, _ = st.columns([1, 1, 3])

        # Re-validate button
        with col_a:
            if st.button("Re-validate", type="primary"):
                # Convert edited DataFrame back to row lists (skip "Errors" col at index 0)
                new_rows = []
                for _, df_row in edited_df.iterrows():
                    new_row = []
                    for ci in range(TOTAL_COLUMNS):
                        val = df_row.iloc[ci + 1]  # +1 to skip "Errors" column
                        if pd.isna(val) or (isinstance(val, str) and val.strip() == ""):
                            new_row.append(None)
                        else:
                            new_row.append(val)
                    new_rows.append(new_row)
                # Filter out completely empty rows
                new_rows = [r for r in new_rows if not is_row_empty(r)]
                new_errors = validate_rows(new_rows, selected_columns)
                st.session_state.sheet_errors[selected_sheet] = new_errors
                st.session_state.sheet_validated_cols[selected_sheet] = selected_columns
                st.session_state.sheet_edits[selected_sheet] = new_rows
                st.rerun()

        # Download button — merges edits from ALL sheets
        with col_b:
            output = io.BytesIO()
            wb_out = openpyxl.Workbook()
            # Remove default sheet; we'll create one per source sheet
            wb_out.remove(wb_out.active)

            for sname in sheet_names:
                ws_out = wb_out.create_sheet(title=sname)
                # Get headers for this sheet
                s_headers = st.session_state.sheet_headers.get(sname)
                if s_headers is None:
                    # Read original headers if not yet loaded
                    try:
                        s_headers, s_orig_rows = read_excel(
                            io.BytesIO(file_bytes), sheet_name=sname
                        )
                        st.session_state.sheet_headers[sname] = s_headers
                    except Exception:
                        continue
                else:
                    _, s_orig_rows = read_excel(
                        io.BytesIO(file_bytes), sheet_name=sname
                    )

                # Use edited rows if available, otherwise original
                s_rows = st.session_state.sheet_edits.get(sname, s_orig_rows)

                # Write headers
                for ci, h in enumerate(s_headers):
                    ws_out.cell(row=1, column=ci + 1, value=h)
                # Write data
                for ri, row in enumerate(s_rows):
                    for ci in range(TOTAL_COLUMNS):
                        val = row[ci] if ci < len(row) else None
                        if isinstance(val, str):
                            try:
                                val = int(val)
                            except ValueError:
                                try:
                                    val = float(val)
                                except ValueError:
                                    pass
                        ws_out.cell(row=ri + 2, column=ci + 1, value=val)

            wb_out.save(output)
            st.download_button(
                label="Download Corrected File",
                data=output.getvalue(),
                file_name=f"corrected_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
